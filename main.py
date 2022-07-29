from pickle import FALSE
from timeit import repeat
from tkinter.tix import Tree
import pandas as pd
from PySimpleGUI import PySimpleGUI as sg
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.command import Command
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common import alert
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.utils import ChromeType

import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.exceptions import *
import re
import csv
import os

class Erro_Pausar_Finalizar: # Essa classe tem o objetivo de ser uma forma para pausar/finalizar a extração do OLX Ninja
    def finalizar(qntdAluguel, qntdAlugueleVenda, qntdForadaÁreadeAtuação):
        horaAtual = time.strftime('%H:%M:%S')
        print(f'\n✔ Filtração Completa! {horaAtual}\n')
        print(f'➙ CEPS | Aluguel: {qntdAluguel}')
        print(f'➙ CEPS | Aluguel e Venda: {qntdAlugueleVenda}')
        print(f'➙ CEPS | Fora da área de atuação: {qntdForadaÁreadeAtuação}')


class MEGAConsultor:
    def __init__(self):
        self.softwareVersion = 'MEGAConsultor - V1.0'
        self.numeroAtual = 0
        self.openChrome = MEGAConsultor.openChrome(self)
        self.interface = MEGAConsultor.interface(self)
        self.Loop = MEGAConsultor.Loop(self)
        
    # def copyInformation(self):
    #     None
        
    def openChrome(self):
        self.user = os.path.expanduser('~')
        if not os.path.exists(fr"{self.user}\Documents\Contas - PONTE Tech"):
            os.makedirs(fr"{self.user}\Documents\Contas - PONTE Tech")
        os.chdir(fr"{self.user}\Documents\Contas - PONTE Tech")
        CHROME_PROFILE_PATH = fr"user-data-dir={self.user}\Documents\Contas - PONTE Tech\ConsultorQA_login"
        chrome_options = webdriver.ChromeOptions(); 
        user_agent= 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_4) AppleWebKit/537.3'
        chrome_options.add_argument('user_agent='+user_agent)
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_argument('disable-notifications')
        chrome_options.add_argument('disable-geolocation')
        chrome_options.add_argument('disable-media-stream')
        chrome_options.add_argument(CHROME_PROFILE_PATH)
        caps = chrome_options.to_capabilities()
        self.browser = webdriver.Chrome(options=chrome_options, desired_capabilities=caps, service=Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()))
        self.browser.maximize_window()
        self.browser.execute_script("return navigator.userAgent;")
        self.browser.get('https://consultor.quintoandar.com.br/auth/login?redirect=%2Fhome%3Flogin%3Dtrue')
        while self.browser.current_url != 'https://consultor.quintoandar.com.br/painel':
            repeat
            
        buttton_registerRealEstate = WebDriverWait(self.browser, 40).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Cadastrar Imóveis']"))).click()
        button_locate = WebDriverWait(self.browser, 40).until(EC.presence_of_element_located((By.XPATH, "//li[2]/button"))).click()
        button_apartament = WebDriverWait(self.browser, 40).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Apartamento']"))).click()
        button_CEP = WebDriverWait(self.browser, 4).until(EC.presence_of_element_located((By.XPATH, "//input[@value='cep']"))).click()

        

    def interface(self):
        sg.theme('Reddit')
        
        layout = [
            [sg.Text('MEGAConsultor', font=('Helvetica', 16, 'bold'), justification='center')],
            [sg.Output(size=(60, 10))],
            [sg.Button('Escolher arquivo', key='-ESCOLHER_EXCEL-'), sg.Button('Iniciar filtração', key='-FILTRAÇÃO-')],
        ]

        créditos = [
            [sg.Text('[ PONTE - SOFTWARES ] ©', text_color='blue')]
        ]

        container = [
            [sg.Column(layout, vertical_alignment='center', element_justification='c')],
            [sg.Column(créditos, vertical_alignment='bottom', element_justification='c')]
        ]

        return sg.Window(self.softwareVersion, container, enable_close_attempted_event=True, grab_anywhere=True, element_justification='center', ttk_theme='clam', use_ttk_buttons=True, size=(530, 300), finalize=True)
    
    def lerExcel(self):
        try:
            tabelaOriginal = pd.read_excel(self.caminho_Excel)
        except ValueError:
            return None
        self.col_CEPS = tabelaOriginal['CEP']
        self.col_nomes = tabelaOriginal['Nome do Proprietário']
        self.col_whatsapps = tabelaOriginal['Número de WhatsApp']
        try:
            self.col_endereço_bairro = tabelaOriginal['Bairro/Município do Imóvel']
        except:
            self.col_endereço_bairro = tabelaOriginal['Endereço do Imóvel']
        self.col_links = tabelaOriginal['Link do Anúncio']
        self.qntdAluguel = 0
        self.qntdAlugueleVenda = 0
        self.qntdForadaÁreadeAtuação = 0
        print(f'Documento Excel selecionado, clique agora em "Iniciar Filtração".')

    def conferir_areaAtuação(self):
        self.modo = None
        CEP = self.browser.find_element(By.XPATH, "//input[@name='zipCode']")
        CEP.click
        CEP.send_keys((Keys.BACK_SPACE)*8)
        jáCarregou = True
        try:
            self.CEPAtual = self.col_CEPS.iloc[self.numeroAtual]
        except IndexError:
            return Erro_Pausar_Finalizar.finalizar(self.qntdAluguel, self.qntdAlugueleVenda, self.qntdForadaÁreadeAtuação)

        except AttributeError:
            print('Nenhum documento Excel selecionado!\n')
            return None

        tamanho = len(str(self.CEPAtual))
        if tamanho == 7:
            self.CEPAtual = f'0{self.CEPAtual}'

        print('\n➙ CEP:', self.CEPAtual)
        CEP.send_keys(str(self.CEPAtual))

        while jáCarregou == True:
            try:
                carregamento = self.browser.find_element(By.CLASS_NAME, 'MuiCircularProgress-svg')
                jáCarregou = True
            except: 
                jáCarregou = False
                repeat

        # Condições de Erro ou Fora da Área de Atuação
        foradaÁrea_confirmação = False
        erronaplataforma_confirmação = False
        
        try:
            foradaÁrea = WebDriverWait(self.browser, 1).until(EC.presence_of_element_located((By.XPATH, "//p[text()='Esse imóvel está fora da nossa área de atuação']")))
            foradaÁrea_confirmação = True
        except:
            None

        try:
            erronaplataforma = WebDriverWait(self.browser, 1).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Tente fazer a busca novamente']"))) 
            erronaplataforma_confirmação = True
        except:
            None

        if foradaÁrea_confirmação == True:
            print('✖ Fora da área de atuação!')
            self.numeroAtual += 1
            self.qntdForadaÁreadeAtuação += 1
            return MEGAConsultor.conferir_areaAtuação(self)
        
        if erronaplataforma_confirmação == True:
            print('✖ Erro no CEP informado!')
            self.numeroAtual += 1
            self.qntdForadaÁreadeAtuação += 1
            return MEGAConsultor.conferir_areaAtuação(self)
        try:
            venda = self.browser.find_element(By.XPATH, "//span[text()='fora']")
            self.modo = 'Aluguel'
            self.qntdAluguel += 1
            print('✔ Aluguel')           
        except:
            self.modo = 'Aluguel/Venda'
            self.qntdAlugueleVenda += 1
            print('✔ Aluguel e Venda')
        return MEGAConsultor.cadastrarnaPlanilha(self)

    def cadastrarnaPlanilha(self):
        try: # Se já existir um Excel com o nome do arquivo atual, as informações serão acrescentar.
            excel = openpyxl.load_workbook(f'{self.caminho_Excel[:-5]} - Filtrado MEGAConsultor.xlsx')
        except FileNotFoundError:
            excel = Workbook()
            fontePadrao = Font(name='Arial', size=14, color='00FFFFFF', bold=True)
            preechimentodaCelula = PatternFill('solid', fgColor='00008000')

            folhadoExcel = excel['Sheet']
            nomeExcel = folhadoExcel['A1']
            nomeExcel.value = 'Nome do Proprietário'

            numeroExcel = folhadoExcel['B1']
            numeroExcel.value = 'Número de WhatsApp'

            bairroExcel = folhadoExcel['C1']
            bairroExcel.value = 'Bairro/Município do Imóvel'

            cepExcel = folhadoExcel['D1']
            cepExcel.value = 'CEP'

            aluguel_venda = folhadoExcel['E1']
            aluguel_venda.value = 'Aluguel ou Venda'

            linkExcel = folhadoExcel['F1']
            linkExcel.value = 'Link do Anúncio'

            nomeExcel.font = fontePadrao
            numeroExcel.font = fontePadrao
            bairroExcel.font = fontePadrao
            cepExcel.font = fontePadrao
            linkExcel.font = fontePadrao
            aluguel_venda.font = fontePadrao

            nomeExcel.fill = preechimentodaCelula
            numeroExcel.fill = preechimentodaCelula
            bairroExcel.fill = preechimentodaCelula
            cepExcel.fill = preechimentodaCelula
            linkExcel.fill = preechimentodaCelula
            aluguel_venda.fill = preechimentodaCelula

            nomeExcel.alignment = Alignment(horizontal="center", vertical="center")
            numeroExcel.alignment = Alignment(horizontal="center", vertical="center")
            bairroExcel.alignment = Alignment(horizontal="center", vertical="center")
            cepExcel.alignment = Alignment(horizontal="center", vertical="center")
            linkExcel.alignment = Alignment(horizontal="center", vertical="center")
            aluguel_venda.alignment = Alignment(horizontal="center", vertical="center")

            folhadoExcel.column_dimensions['A'].width = 40
            folhadoExcel.column_dimensions['B'].width = 40
            folhadoExcel.column_dimensions['C'].width = 55 
            folhadoExcel.column_dimensions['D'].width = 15
            folhadoExcel.column_dimensions['E'].width = 30
            folhadoExcel.column_dimensions['F'].width = 30

        folhadoExcel = excel['Sheet']
        
        try:
            folhadoExcel.append([self.col_nomes.iloc[self.numeroAtual], self.col_whatsapps.iloc[self.numeroAtual], self.col_endereço_bairro.iloc[self.numeroAtual], self.CEPAtual, self.modo, self.col_links.iloc[self.numeroAtual]])
        except IndexError:
            return Erro_Pausar_Finalizar.finalizar(self.qntdAluguel, self.qntdAlugueleVenda, self.qntdForadaÁreadeAtuação)

        excel.save(f'{self.caminho_Excel[:-5]} - Filtrado MEGAConsultor.xlsx')

        self.numeroAtual += 1
        return MEGAConsultor.conferir_areaAtuação(self)
                
    def Loop(self):
        self.estado = None
        while True:
            event, values = self.interface.Read()
            if event == '-ESCOLHER_EXCEL-': 
                self.caminho_Excel = sg.popup_get_file(title=self.softwareVersion, message='Escolha o documento Excel a ser filtrado.', file_types=(('Apenas EXCEL', '*.xlsx'), ('Todos arquivos', '*.*')), save_as=False, multiple_files=True, grab_anywhere=True)
                self.numeroAtual = 0
                MEGAConsultor.lerExcel(MEGAConsultor)
            
            if event == '-FILTRAÇÃO-':
                self.interface.perform_long_operation(lambda: MEGAConsultor.conferir_areaAtuação(self), '')
                
            if event == sg.WIN_CLOSED:
                break

MEGAConsultor.__init__(MEGAConsultor)
